import pathlib
from pathlib import Path

import colorama
from colorama import Fore, Back
colorama.init(autoreset=True)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

option = ["a", "c"]
location = input(Fore.BLUE + "Absolute path or Current directory? (a/c): ")
flag = 0

for letter in option:
    if location.lower() == "c":
        flag = 1
        break

if flag == 1:
        print(Fore.BLACK + Back.YELLOW + " FILE MUST BE UPLOADED BEFORE STARTING ")
        filename = input(Fore.BLUE + "Enter file name: ")

        path = Path()
        for exel in path.glob(filename):
            file = path.exists()
            if file == True:

                rename = input(Fore.BLUE + "Rename your file: ")
                wb = xl.load_workbook(filename)
                sheet = wb['Sheet1']

                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row, 3)
                    corrected_price = cell.value * 0.9
                    corrected_price_cell = sheet.cell(row, 4)
                    corrected_price_cell.value = corrected_price

                values = Reference(sheet,
                                   min_row=2,
                                   max_row=sheet.max_row,
                                   min_col=4,
                                   max_col=4)

                chart = BarChart()
                chart.add_data(values)
                sheet.add_chart(chart, 'a6')

                wb.save(rename)
                print(Fore.BLACK + Back.GREEN + " SPREADSHEET SUCCESSFULLY UPDATED ")
                print(Fore.BLACK + Back.GREEN + " OPEN IN FILE EXPLORER TO SEE THE CHANGES ")

for letter in option:
    if location.lower() == "a":
        flag = 2
        break

if flag == 2:
        print(Fore.BLACK + Back.YELLOW + " ENTER CORRECT FILE PATH ")
        absolute = input(Fore.BLUE + "Enter file path: ")

        path_rename = input(Fore.BLUE + "Rename your file: ")
        wb = xl.load_workbook(absolute)
        sheet = wb['Sheet1']

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            corrected_price = cell.value * 0.9
            corrected_price_cell = sheet.cell(row, 4)
            corrected_price_cell.value = corrected_price

        values = Reference(sheet,
                            min_row=2,
                            max_row=sheet.max_row,
                            min_col=4,
                            max_col=4)

        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'a6')

        wb.save(path_rename)
        print(Fore.BLACK + Back.GREEN + " SPREADSHEET SUCCESSFULLY UPDATED ")
        print(Fore.BLACK + Back.GREEN + " OPEN IN FILE EXPLORER TO SEE THE CHANGES ")

for letter in option:
    if location not in option:
        print(Fore.RED + "FILE NOT FOUND")
